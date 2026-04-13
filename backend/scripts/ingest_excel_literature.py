#!/usr/bin/env python3
"""
单独将「文献 Excel」按行写入 Milvus：每行一条向量，Abstract 作为 text（不做三层滑动切分，但 chunk_level 与 RAG 叶子层级一致，否则检索被 chunk_level 过滤掉）。

表头（须与首行一致，列名可前后空格）：
  Title | Journal Abbreviation | Publication Date | PMID | Pubmed Web | DOI | PMC |
  Abstract | Citation Counts | JournalTitle | category | if_2024 | _source_file

写入 Milvus 的 text：Abstract
动态字段 meta（JSON 字符串）：title, publication_date, journal_title, category, pubmed_web
其余列不落库。

用法（在仓库根目录，默认写死路径与 brief，可直接运行）：
  uv run python backend/scripts/ingest_excel_literature.py
  uv run python backend/scripts/ingest_excel_literature.py --file /其它路径.xlsx --kb-tier detailed

不修改项目内其它模块，仅依赖现有 embedding / milvus_client。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPTS_DIR.parent
ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))
os.chdir(BACKEND_DIR)

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

# 与 backend/rag_utils.py 中 retrieve_documents 的 filter「chunk_level == LEAF_RETRIEVE_LEVEL」一致
LEAF_RETRIEVE_LEVEL = int(os.getenv("LEAF_RETRIEVE_LEVEL", "3"))

# =============================================================================
# 写死：默认入库 Excel 与 KB 分层（相对项目根 data/）
# =============================================================================
EXCEL_PATH: Path = (
    ROOT
    / "data"
    / "abstract and mini golden paper_quick"
    / "EN_Pummed Abstract_NPC Nurse_final.xlsx"
)
KB_TIER: str = "detailed"

# Milvus `text` 字段：DataType.VARCHAR, max_length=2400（见 milvus_client.init_collection）。
# 该上限按 UTF-8 字节计，不是 Python len(str) 的字符个数；中文等宽字符易触发「已按 2000 字符截断仍超限制」。
MILVUS_TEXT_MAX_BYTES = 2400
BATCH_SIZE = 32


def _truncate_milvus_text_field(s: str, max_bytes: int = MILVUS_TEXT_MAX_BYTES) -> tuple[str, bool]:
    """
    将 Abstract 截断为不超过 max_bytes 个 UTF-8 字节，避免插入 Milvus 报错。
    返回 (截断后的文本, 是否发生过截断)。
    """
    if not s:
        return "", False
    raw = s.encode("utf-8")
    if len(raw) <= max_bytes:
        return s, False
    cut = raw[:max_bytes]
    while cut:
        try:
            return cut.decode("utf-8"), True
        except UnicodeDecodeError:
            cut = cut[:-1]
    return "", True

# 期望表头（用于校验与列索引）
EXPECTED_HEADERS = [
    "Title",
    "Journal Abbreviation",
    "Publication Date",
    "PMID",
    "Pubmed Web",
    "DOI",
    "PMC",
    "Abstract",
    "Citation Counts",
    "JournalTitle",
    "category",
    "if_2024",
]


def _norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip()


def _build_header_map(header_row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key:
            mapping[key] = i
    return mapping


def _cell(row: tuple, col_map: dict[str, int], name: str) -> str:
    idx = col_map.get(name)
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def _safe_filename_part(s: str, max_len: int = 80) -> str:
    s = re.sub(r'[<>:"/\\|?*]', "_", s).strip()
    if not s:
        return "row"
    return s[:max_len]


def load_rows(excel_path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        wb.close()
        raise SystemExit("Excel 为空")

    col_map = _build_header_map(header_row)
    missing = [h for h in EXPECTED_HEADERS if h not in col_map]
    if missing:
        wb.close()
        raise SystemExit(
            f"表头缺少列: {missing}\n当前识别到的列: {list(col_map.keys())}"
        )

    rows_out: list[dict] = []
    chunk_seq = 0
    for row_idx, row in enumerate(it, start=2):
        abstract = _cell(row, col_map, "Abstract")
        if not abstract:
            continue
        title = _cell(row, col_map, "Title")
        pub_date = _cell(row, col_map, "Publication Date")
        journal_title = _cell(row, col_map, "JournalTitle")
        category = _cell(row, col_map, "category")
        pubmed_web = _cell(row, col_map, "Pubmed Web")
        source_file = _cell(row, col_map, "_source_file")

        meta = {
            "title": title,
            "publication_date": pub_date,
            "journal_title": journal_title,
            "category": category,
            "pubmed_web": pubmed_web,
        }
        text, truncated = _truncate_milvus_text_field(abstract, MILVUS_TEXT_MAX_BYTES)
        if truncated:
            print(
                f"[WARN] 第 {row_idx} 行 Abstract 超过 Milvus text 上限（{MILVUS_TEXT_MAX_BYTES} UTF-8 字节），已截断后入库",
                file=sys.stderr,
            )

        stem = excel_path.stem
        base_fn = Path(source_file).name if source_file else f"{stem}_{row_idx}_{title or 'untitled'}"
        fname = _safe_filename_part(base_fn)[:200]

        lvl = LEAF_RETRIEVE_LEVEL
        cid = f"{stem}::row{row_idx}::l{lvl}::0"
        rows_out.append(
            {
                "text": text,
                "meta": json.dumps(meta, ensure_ascii=False),
                "filename": fname,
                "file_path": str(excel_path.resolve()),
                "page_number": row_idx,
                "chunk_idx": chunk_seq,
                "chunk_id": cid,
                "parent_chunk_id": "",
                "root_chunk_id": cid,
                "chunk_level": lvl,
            }
        )
        chunk_seq += 1

    wb.close()
    return rows_out


def main():
    ap = argparse.ArgumentParser(description="Excel 文献表按行写入 Milvus")
    ap.add_argument(
        "--file",
        "-f",
        type=Path,
        default=None,
        help=f"xlsx 文件路径（默认写死: {EXCEL_PATH}）",
    )
    ap.add_argument(
        "--kb-tier",
        default=KB_TIER,
        choices=("brief", "detailed"),
        help=f"写入 brief 或 detailed 集合（默认 {KB_TIER}）",
    )
    args = ap.parse_args()
    excel_path = (args.file if args.file is not None else EXCEL_PATH).resolve()
    if not excel_path.is_file():
        raise SystemExit(f"文件不存在: {excel_path}")

    from embedding import embedding_service
    from milvus_client import MilvusManager

    documents = load_rows(excel_path)
    if not documents:
        print("没有可入库行（Abstract 为空的行已跳过）")
        return

    milvus = MilvusManager()
    milvus.init_collection(kb_tier=args.kb_tier)

    texts = [d["text"] for d in documents]
    embedding_service.increment_add_documents(texts)

    total = len(documents)
    file_type = "Excel"
    for i in range(0, total, BATCH_SIZE):
        batch = documents[i : i + BATCH_SIZE]
        tbatch = [d["text"] for d in batch]
        dense_embeddings, sparse_embeddings = embedding_service.get_all_embeddings(tbatch)

        insert_data = []
        for doc, dense_emb, sparse_emb in zip(batch, dense_embeddings, sparse_embeddings):
            insert_data.append(
                {
                    "dense_embedding": dense_emb,
                    "sparse_embedding": sparse_emb,
                    "text": doc["text"],
                    "filename": doc["filename"],
                    "file_type": file_type,
                    "file_path": doc["file_path"],
                    "page_number": int(doc["page_number"]),
                    "chunk_idx": int(doc["chunk_idx"]),
                    "chunk_id": doc["chunk_id"],
                    "parent_chunk_id": doc["parent_chunk_id"],
                    "root_chunk_id": doc["root_chunk_id"],
                    "chunk_level": int(doc["chunk_level"]),
                    "meta": doc["meta"],
                }
            )
        milvus.insert(insert_data, kb_tier=args.kb_tier)
        print(f"已写入 {min(i + BATCH_SIZE, total)} / {total}")

    print(f"完成：共 {total} 条写入 Milvus（kb_tier={args.kb_tier}）")


if __name__ == "__main__":
    main()
